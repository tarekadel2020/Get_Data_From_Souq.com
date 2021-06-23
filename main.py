from tkinter import *
from tkinter import messagebox , filedialog
import tkinter as tk
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re

## https://www.youtube.com/watch?v=7YS6YDQKFh0
## https://www.youtube.com/watch?v=t51bT7WbeCM&ab_channel=ProgrammingKnowledgeProgrammingKnowledge

def souq():
    try:
        wb = load_workbook(Enter_Path.get())
    except:
        messagebox.showinfo("Erorr", "Please Enter Path")

    ws = wb.active

    try:
        x = int(Enter_Cell.get())+1
    except:
        messagebox.showinfo("Erorr", "Please Enter End Cell")

##print(ws['A20'].value)

    for cal in range(2,x):
        if ws["A"+ str(cal)].value == None:
            print("error")
        else:
            print(ws["A"+ str(cal)].value)
            zz = ws["A" + str(cal)].value
            r = requests.get(zz)
            soup = BeautifulSoup(r.content, "html.parser")
            for souq_title in soup.findAll('head') :
                print(souq_title.text.strip())
                ws["B" + str(cal)].value = souq_title.text.strip()
                try:
                    wb.save(Enter_Path.get())
                except:
                    messagebox.showinfo("Erorr", "Please Close The Excel")

            for souq_name in soup.findAll('h1'):
                #print(souq_name.text)
                ws["C" + str(cal)].value = souq_name.text
                try:
                    wb.save(Enter_Path.get())
                except:
                    messagebox.showinfo("Erorr", "Please Close The Excel")

            for souq_price in soup.findAll('div', class_='columns large-8 medium-8 small-7') :
                #print(souq_price.text.strip())
                ws["D" + str(cal)].value = souq_price.text.strip()
                try:
                    wb.save(Enter_Path.get())
                except:
                    messagebox.showinfo("Erorr", "Please Close The Excel")

            for souq_qte in soup.findAll('div', class_='unit-labels'):
                #print(souq_qte.text.strip())
                ws["E" + str(cal)].value = souq_qte.text.strip()
                try:
                    wb.save(Enter_Path.get())
                except:
                    messagebox.showinfo("Erorr", "Please Close The Excel")

            for url_image in soup.findAll('img', attrs={'src':re.compile('jpg')}):
                #print(url_image.get('src'))
                ws["F" + str(cal)].value = url_image.get('src')
                try:
                    wb.save(Enter_Path.get())
                except:
                    messagebox.showinfo("Erorr", "Please Close The Excel")
    messagebox.showinfo("Done", "Done")



def select():
    root = tk.Tk()
    root.withdraw()
    file = str(filedialog.askopenfilename())
    Enter_Path.insert(0,file)


top = Tk()
top.title("Souq.com V_0.0.1")

width = 500
hight = 300

screen_width = top.winfo_screenwidth()
screen_hight = top.winfo_screenheight()

x_d = int((screen_width/2) - (width/2))
y_d =  int((screen_hight/2) - (hight/2))


Path = Label(text="Path")
Path.pack()

Enter_Path = Entry(top , width=70)
Enter_Path.pack()

but_1 = Button(text="Select" , command=select)
but_1.pack()

space = Label(text="  ")
space.pack()

Num = Label(text="Last Cell")
Num.pack()

Enter_Cell = Entry(top , width=10)
Enter_Cell.pack()

but = Button(text="RUN", command=souq)
but.pack()


space = Label(text="  ")
space.pack()

space = Label(text="  ")
space.pack()


space = Label(text="  ")
space.pack()

space = Label(text="  ")
space.pack()

space = Label(text="Copyright (C) 2021")
space.pack()

space = Label(text="Powered By : Tarek Adel")
space.pack()
top.geometry(f'{width}x{hight}+{x_d}+{y_d}')
top.mainloop()
